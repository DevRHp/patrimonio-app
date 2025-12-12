
import sys
import unittest.mock
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# Mock UI libraries
sys.modules['customtkinter'] = unittest.mock.MagicMock()
sys.modules['tkinter'] = unittest.mock.MagicMock()
sys.modules['tkinter.filedialog'] = unittest.mock.MagicMock()
sys.modules['tkinter.messagebox'] = unittest.mock.MagicMock()
sys.modules['tkinter.ttk'] = unittest.mock.MagicMock()
sys.modules['PIL'] = unittest.mock.MagicMock()
sys.modules['PIL.Image'] = unittest.mock.MagicMock()

# Create dummy input files
if not os.path.exists("test_input"):
    os.makedirs("test_input")

def create_dummy_excel(filename, data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in data:
        ws.append(row)
    wb.save(filename)

file_a = os.path.abspath("test_input/test_A.xlsx")
file_b = os.path.abspath("test_input/test_B.xlsx")

create_dummy_excel(file_a, [["Item", "Code"], ["Chair", "123"], ["Desk", "456"]])
create_dummy_excel(file_b, [["Item", "Code"], ["Monitor", "789"], ["Mouse", "000"]])

# Standalone logic mirroring the changes in main_fe.py
def process_all_files_standalone(filepaths, scanned_codes):
    print("Starting standalone processing...")
    output_folder = "Relatorios_Gerados"
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    global_found_items = set()
    files_processed_count = 0

    master_wb = Workbook()
    if master_wb.active:
        master_wb.remove(master_wb.active)

    try:
        for filepath in filepaths:
            print(f"Processing {filepath}")
            wb = load_workbook(filepath)
            base_name = os.path.splitext(os.path.basename(filepath))[0][:30]

            for sheet in wb.worksheets:
                sheet_title = base_name if len(wb.worksheets) == 1 else f"{base_name}_{sheet.title}"[:30]
                ws_master = master_wb.create_sheet(title=sheet_title)

                for row in sheet.iter_rows():
                    for cell in row:
                        ws_master.cell(row=cell.row, column=cell.column, value=cell.value)

                for row in ws_master.iter_rows():
                    match_in_row = False
                    for cell in row:
                        if cell.value is not None:
                            cell_val = str(cell.value).strip()
                            if cell_val in scanned_codes:
                                match_in_row = True
                                global_found_items.add(cell_val)
                    
                    if match_in_row:
                        for cell in row:
                            cell.fill = green_fill

            files_processed_count += 1

        master_save_path = os.path.join(output_folder, "RELATORIO_GERAL_PROCESSADO.xlsx")
        print(f"Saving master to {master_save_path}")
        master_wb.save(master_save_path)

        not_found_items = scanned_codes - global_found_items
        if not_found_items:
            wb_extras = Workbook()
            ws_extras = wb_extras.active
            ws_extras.title = "Itens Nao Encontrados"
            ws_extras.append(["Codigo Lido", "Status"])
            for item in not_found_items:
                ws_extras.append([item, "NAO ENCONTRADO"])
            
            extras_path = os.path.join(output_folder, "RELATORIO_ITENS_SOBRANDO.xlsx")
            wb_extras.save(extras_path)
            print(f"Saved extras to {extras_path}")
        
    except Exception as e:
        print(f"Error in standalone: {e}")
        import traceback
        traceback.print_exc()

# Run
try:
    process_all_files_standalone([file_a, file_b], {"123", "789"})
    print("Process finished.")
except Exception as e:
    print(f"Process crashed: {e}")
    import traceback
    traceback.print_exc()

# Verify output
output_folder = os.path.abspath("Relatorios_Gerados")
output_file = os.path.join(output_folder, "RELATORIO_GERAL_PROCESSADO.xlsx")

if os.path.exists(output_file):
    print(f"Output file found: {output_file}")
    wb = load_workbook(output_file)
    print(f"Sheet names: {wb.sheetnames}")
    
    ws_a = wb[wb.sheetnames[0]] 
    c123 = ws_a['B2']
    print(f"Cell B2 value: {c123.value}")
    if c123.fill and c123.fill.start_color.index == "00FF00":
         print("Cell B2 is GREEN (Correct)")
    else:
         print(f"Cell B2 fill: {c123.fill.start_color.index if c123.fill else 'None'}")

    c456 = ws_a['B3']
    print(f"Cell B3 value: {c456.value}")
    if not c456.fill or c456.fill.start_color.index == "00000000":
        print("Cell B3 is NOT GREEN (Correct)")
    else:
        print(f"Cell B3 fill: {c456.fill.start_color.index if c456.fill else 'None'}")
else:
    print("Output file NOT found!")
