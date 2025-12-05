import os
import zipfile
import io
from flask import Flask, render_template, request, send_file, jsonify, session
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from copy import copy

app = Flask(__name__)
app.secret_key = 'super_secret_key_sesi_sorocaba' # Change this in production!

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
REPORTS_FOLDER = os.path.join(UPLOAD_FOLDER, 'reports')
os.makedirs(REPORTS_FOLDER, exist_ok=True)

# Global variable to store the path of the uploaded master file
# In a production app, this should be handled per session or database
MASTER_FILE_PATH = None
POSSIBLE_MASTER_PATH = os.path.join(UPLOAD_FOLDER, 'master_spreadsheet.xlsx')
if os.path.exists(POSSIBLE_MASTER_PATH):
    MASTER_FILE_PATH = POSSIBLE_MASTER_PATH

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['POST'])
def login():
    data = request.json
    email = data.get('email')
    password = data.get('password')

    if email == 'admin@gmail.com' and password == 'admin123':
        session['is_admin'] = True
        return jsonify({'message': 'Login realizado com sucesso', 'success': True})
    else:
        return jsonify({'error': 'Credenciais inválidas', 'success': False}), 401

@app.route('/logout', methods=['POST'])
def logout():
    session.pop('is_admin', None)
    return jsonify({'message': 'Logout realizado com sucesso'})

@app.route('/check_auth', methods=['GET'])
def check_auth():
    return jsonify({'is_admin': session.get('is_admin', False)})

@app.route('/upload_master', methods=['POST'])
def upload_master():
    if not session.get('is_admin'):
        return jsonify({'error': 'Acesso negado. Requer privilégios de administrador.'}), 403

    global MASTER_FILE_PATH
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'master_spreadsheet.xlsx')
        file.save(filepath)
        MASTER_FILE_PATH = filepath
        
        # Load workbook to get sheet names and extract room names
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            rooms_data = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                room_display_name = sheet_name # Default to sheet name
                
                # Search for "Denominação"
                found_header = False
                # Limit search to first 20 rows and 20 columns for performance
                for row in ws.iter_rows(min_row=1, max_row=20, max_col=20):
                    for cell in row:
                        if cell.value and str(cell.value).strip() == "Denominação":
                            # Found header, get value from cell below
                            target_row = cell.row + 1
                            target_col = cell.column
                            # In read_only mode, we can access cell by coordinate
                            try:
                                val = ws.cell(row=target_row, column=target_col).value
                                if val:
                                    room_display_name = str(val).strip()
                                    found_header = True
                            except:
                                pass
                            break
                    if found_header:
                        break
                
                rooms_data.append({'id': sheet_name, 'name': room_display_name})

            wb.close()
            return jsonify({'message': 'File uploaded successfully', 'rooms': rooms_data})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

@app.route('/delete_master', methods=['POST'])
def delete_master():
    if not session.get('is_admin'):
        return jsonify({'error': 'Acesso negado. Requer privilégios de administrador.'}), 403

    global MASTER_FILE_PATH
    if MASTER_FILE_PATH and os.path.exists(MASTER_FILE_PATH):
        try:
            os.remove(MASTER_FILE_PATH)
            MASTER_FILE_PATH = None
            return jsonify({'message': 'Planilha mãe removida com sucesso'})
        except Exception as e:
            return jsonify({'error': f'Erro ao remover arquivo: {str(e)}'}), 500
    else:
        MASTER_FILE_PATH = None # Ensure it's None even if file didn't exist
        return jsonify({'message': 'Nenhuma planilha mãe para remover'})

@app.route('/check_master', methods=['GET'])
def check_master():
    global MASTER_FILE_PATH
    if MASTER_FILE_PATH and os.path.exists(MASTER_FILE_PATH):
        # You might want to return more info like upload date in a real app
        return jsonify({'exists': True, 'filename': os.path.basename(MASTER_FILE_PATH)})
    else:
        return jsonify({'exists': False})

@app.route('/get_rooms', methods=['GET'])
def get_rooms():
    global MASTER_FILE_PATH
    if not MASTER_FILE_PATH or not os.path.exists(MASTER_FILE_PATH):
        return jsonify({'error': 'Planilha mãe não encontrada'}), 404
        
    try:
        wb = load_workbook(MASTER_FILE_PATH, read_only=True, data_only=True)
        rooms_data = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            room_display_name = sheet_name # Default to sheet name
            
            # Search for "Denominação"
            found_header = False
            # Limit search to first 20 rows and 20 columns for performance
            # Note: iter_rows in read_only might be slower if scanning whole sheet, 
            # but here we limit max_row/col.
            # However, reusing the logic from upload_master is best.
            # Actually, upload_master logic is good.
            
            for row in ws.iter_rows(min_row=1, max_row=20, max_col=20):
                for cell in row:
                    if cell.value and str(cell.value).strip() == "Denominação":
                        # Found header, get value from cell below
                        target_row = cell.row + 1
                        target_col = cell.column
                        try:
                            # In read_only mode, cell access by coordinate is ok
                            # But avoiding ws.cell() in read_only if possible.
                            # Just iterating again is safer or using the offset?
                            # ws.cell(row=target_row, column=target_col).value works in recent openpyxl
                            val = ws.cell(row=target_row, column=target_col).value
                            if val:
                                room_display_name = str(val).strip()
                                found_header = True
                        except:
                            pass
                        break
                if found_header:
                    break
            
            rooms_data.append({'id': sheet_name, 'name': room_display_name})

        wb.close()
        return jsonify({'rooms': rooms_data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/list_reports', methods=['GET'])
def list_reports():
    try:
        reports = []
        if os.path.exists(REPORTS_FOLDER):
            for filename in os.listdir(REPORTS_FOLDER):
                if filename.endswith('.zip'):
                    filepath = os.path.join(REPORTS_FOLDER, filename)
                    stats = os.stat(filepath)
                    reports.append({
                        'filename': filename,
                        'created_at': stats.st_mtime,
                        'size': stats.st_size
                    })
        # Sort by newest first
        reports.sort(key=lambda x: x['created_at'], reverse=True)
        return jsonify({'reports': reports})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get_report/<path:filename>', methods=['GET'])
def get_report(filename):
    try:
        return send_file(os.path.join(REPORTS_FOLDER, filename), as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@app.route('/verify', methods=['POST'])
def verify():
    global MASTER_FILE_PATH
    if not MASTER_FILE_PATH or not os.path.exists(MASTER_FILE_PATH):
        return jsonify({'error': 'Planilha mãe não encontrada. Por favor, faça o upload novamente.'}), 400

    data = request.json
    analyst_name = data.get('analyst_name', 'Analista')
    selected_room = data.get('room_name')
    scanned_codes_raw = data.get('scanned_codes', '')
    
    # Normalize scanned codes
    scanned_codes = set(code.strip() for code in scanned_codes_raw.splitlines() if code.strip())

    if not selected_room:
        return jsonify({'error': 'Nenhuma sala selecionada'}), 400

    try:
        # Load the master workbook
        wb = load_workbook(MASTER_FILE_PATH)
        
        if selected_room not in wb.sheetnames:
             return jsonify({'error': f'Sala "{selected_room}" não encontrada na planilha mãe'}), 400

        source_ws = wb[selected_room]
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        # Identify found items in the target room
        found_in_room_codes = set()
        
        # We need to scan the sheet first to find matches
        # Assuming header is row 1
        for row in source_ws.iter_rows(min_row=2, values_only=True):
            row_values = [str(cell).strip() for cell in row if cell is not None]
            for val in row_values:
                if val in scanned_codes:
                    found_in_room_codes.add(val)
                    break

        # --- 1. Verified Items (Sheet 1) ---
        # Strategy: Copy the original sheet, then highlight found items.
        # This preserves all formatting (merged cells, widths, fonts, etc.)
        verified_ws = wb.copy_worksheet(source_ws)
        verified_ws.title = "Verificados"
        
        # Highlight found items in the new sheet
        for row in verified_ws.iter_rows(min_row=2):
            match_found = False
            for cell in row:
                if cell.value is not None and str(cell.value).strip() in scanned_codes:
                    match_found = True
                    break
            
            if match_found:
                for cell in row:
                    cell.fill = green_fill

        # --- 3. Missing Items (Sheet 3) ---
        # Strategy: Create a new sheet, copy dimensions/header, then copy ONLY missing rows with styles.
        # Deleting rows from a copied sheet is slow/buggy, so copying row-by-row is safer for "filtering".
        missing_ws = wb.create_sheet("Nao Encontrados")
        
        # Copy column dimensions
        for col_name, col_dim in source_ws.column_dimensions.items():
            missing_ws.column_dimensions[col_name].width = col_dim.width

        # Copy Header (Row 1)
        for row in source_ws.iter_rows(min_row=1, max_row=1):
            missing_ws.append([cell.value for cell in row])
            # Copy header styles
            for i, cell in enumerate(row):
                new_cell = missing_ws.cell(row=1, column=i+1)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Copy Missing Rows
        # from copy import copy (Removed local import)
        current_row_idx = 2
        for row in source_ws.iter_rows(min_row=2):
            # Check if this row represents a found item
            is_found = False
            for cell in row:
                if cell.value is not None and str(cell.value).strip() in found_in_room_codes:
                    is_found = True
                    break
            
            if not is_found:
                # Copy this row
                for i, cell in enumerate(row):
                    new_cell = missing_ws.cell(row=current_row_idx, column=i+1, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
                current_row_idx += 1


        # --- 2. Wrong Location Items (Sheet 2) ---
        # Scanned items NOT in target room.
        wrong_location_ws = wb.create_sheet("Local Incorreto")
        wrong_location_ws.append(["Codigo", "Encontrado na Sala", "Deveria estar em"]) 
        # Style the header simply
        for cell in wrong_location_ws[1]:
            cell.font = copy(source_ws.cell(1,1).font) # Try to copy font from source header

        scanned_but_not_in_room = scanned_codes - found_in_room_codes
        
        for code in scanned_but_not_in_room:
            found_location = "Nao encontrado na planilha mae"
            for sheet_name in wb.sheetnames:
                if sheet_name in [selected_room, "Verificados", "Nao Encontrados", "Local Incorreto"]:
                    continue
                
                sheet = wb[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_str_values = [str(v).strip() for v in row if v is not None]
                    if code in row_str_values:
                        found_location = sheet_name
                        break
                if found_location != "Nao encontrado na planilha mae":
                    break
            
            wrong_location_ws.append([code, selected_room, found_location])

        # --- Save and Zip ---
        # We need to save the modified workbook (which contains verified/missing sheets)
        # AND create a new workbook for the "Wrong Location" if we wanted separate files, 
        # BUT the user asked for "3 planilhas dentro de umas pasta" (3 files in a folder/zip).
        # My previous implementation created 3 separate Workbooks.
        # The user's new request "manter o padrão" implies they want the original formatting.
        # So I will extract the 3 sheets I created into separate Workbooks to save them as files.

        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            
            # 1. Verified File
            wb_verified = Workbook()
            wb_verified.remove(wb_verified.active)
            # Copy the "Verificados" sheet from the main wb to this new wb
            # Copying between workbooks is hard in openpyxl.
            # EASIER: Save the main wb as "Verified", delete other sheets?
            # EASIER: Just save the main wb with the new sheets?
            # User said: "gero a planilha fablab_verificada... e quero que gere uma outra planilha... ao total seram 3 planilhas"
            # implying 3 FILES.
            
            # To preserve formatting perfectly, it's best to save the whole WB as the "Verified" file, 
            # but maybe strip other sheets?
            # Let's try to be efficient.
            
            # File 1: Verified
            # We already have 'verified_ws' in 'wb'.
            # Let's delete all other sheets from 'wb' except 'verified_ws' and save it?
            # No, we need 'wb' for the other steps.
            
            # Let's save the whole 'wb' to a temp buffer, reload it 3 times, and delete unwanted sheets?
            # That ensures perfect formatting copy.
            
            # Save current state (with Verificados, Nao Encontrados, Local Incorreto added)
            temp_buffer = io.BytesIO()
            wb.save(temp_buffer)
            temp_buffer.seek(0)
            
            # --- File 1: Verified ---
            wb_v = load_workbook(temp_buffer)
            # Keep only "Verificados"
            for s in wb_v.sheetnames:
                if s != "Verificados":
                    del wb_v[s]
            with io.BytesIO() as f:
                wb_v.save(f)
                zf.writestr(f"{analyst_name}_Verificados.xlsx", f.getvalue())
            
            # --- File 2: Missing ---
            temp_buffer.seek(0)
            wb_m = load_workbook(temp_buffer)
            # Keep only "Nao Encontrados"
            for s in wb_m.sheetnames:
                if s != "Nao Encontrados":
                    del wb_m[s]
            with io.BytesIO() as f:
                wb_m.save(f)
                zf.writestr(f"{analyst_name}_Nao_Encontrados.xlsx", f.getvalue())

            # --- File 3: Wrong Location ---
            temp_buffer.seek(0)
            wb_w = load_workbook(temp_buffer)
            # Keep only "Local Incorreto"
            for s in wb_w.sheetnames:
                if s != "Local Incorreto":
                    del wb_w[s]
            with io.BytesIO() as f:
                wb_w.save(f)
                zf.writestr(f"{analyst_name}_Local_Incorreto.xlsx", f.getvalue())

        memory_file.seek(0)
        

        
        # Save copy to reports folder
        import time
        timestamp = int(time.time())
        report_filename = f"{analyst_name}_{timestamp}_Relatorios.zip"
        report_path = os.path.join(REPORTS_FOLDER, report_filename)
        
        with open(report_path, 'wb') as f:
            f.write(memory_file.getvalue())

        memory_file.seek(0)
        
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'{analyst_name}_Relatorios.zip'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
