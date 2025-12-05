import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import os

class MultiInventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestor de Ativos - Multi-Planilhas [Elite]")
        self.root.geometry("800x600")

        # Variáveis de Estado
        self.filepaths = []  # Lista para vários arquivos
        self.scanned_codes = set() # Todos os códigos lidos pelo scanner
        
        # --- Interface ---
        
        # 1. Área de Seleção de Arquivos
        self.frame_top = tk.Frame(root, pady=15, bg="#e8e8e8")
        self.frame_top.pack(fill=tk.X)
        
        self.btn_load = tk.Button(self.frame_top, text="1. Selecionar Planilhas (Pode escolher várias)", 
                                  command=self.load_files, font=("Arial", 11))
        self.btn_load.pack(pady=5)
        
        self.lbl_files = tk.Label(self.frame_top, text="Nenhum arquivo selecionado", bg="#e8e8e8", fg="#555")
        self.lbl_files.pack()

        # 2. Área de Escaneamento
        self.frame_scan = tk.Frame(root, pady=15)
        self.frame_scan.pack()
        
        tk.Label(self.frame_scan, text="BIPAR CÓDIGO (Foco aqui):", font=("Arial", 12, "bold")).pack()
        
        self.entry_scan = tk.Entry(self.frame_scan, font=("Arial", 24), width=20, justify='center', bg="#fffacd")
        self.entry_scan.pack(pady=10)
        self.entry_scan.bind('<Return>', self.on_scan)
        self.entry_scan.focus_set()

        # 3. Lista Visual
        self.tree_frame = tk.Frame(root)
        self.tree_frame.pack(expand=True, fill='both', padx=20)
        
        # Treeview com barra de rolagem
        self.tree = ttk.Treeview(self.tree_frame, columns=('code',), show='headings', height=8)
        self.tree.heading('code', text='Fila de Processamento (Itens Lidos)')
        self.tree.column('code', anchor='center')
        self.tree.pack(side=tk.LEFT, expand=True, fill='both')
        
        sb = tk.Scrollbar(self.tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        sb.pack(side=tk.RIGHT, fill='y')
        self.tree.configure(yscrollcommand=sb.set)

        self.lbl_count = tk.Label(root, text="Total lido: 0", font=("Arial", 10, "bold"))
        self.lbl_count.pack(pady=5)

        # 4. Botão de Execução
        self.btn_process = tk.Button(root, text="2. PROCESSAR TUDO E GERAR RELATÓRIO DE SOBRAS", 
                                   command=self.process_all_files, 
                                   bg="#28a745", fg="white", font=("Arial", 12, "bold"), height=2)
        self.btn_process.pack(fill=tk.X, padx=20, pady=20)

    def load_files(self):
        # Permite selecionar múltiplos arquivos
        filenames = filedialog.askopenfilenames(title="Selecione as Planilhas", 
                                                filetypes=[("Excel files", "*.xlsx")])
        if filenames:
            self.filepaths = filenames
            qtd = len(filenames)
            self.lbl_files.config(text=f"{qtd} arquivos prontos para análise.", fg="blue")
            self.entry_scan.focus_set()

    def on_scan(self, event):
        code = self.entry_scan.get().strip()
        if code:
            code_normalized = str(code) # Garante formato texto
            
            if code_normalized not in self.scanned_codes:
                self.scanned_codes.add(code_normalized)
                self.tree.insert('', 0, values=(code_normalized,))
                self.lbl_count.config(text=f"Total lido: {len(self.scanned_codes)}")
                
            self.entry_scan.delete(0, tk.END)

    def process_all_files(self):
        if not self.filepaths:
            messagebox.showwarning("Erro", "Selecione as planilhas primeiro!")
            return
        
        if not self.scanned_codes:
            messagebox.showwarning("Erro", "Nenhum código foi bipado!")
            return

        # Cria pasta de saída para não bagunçar os originais
        output_folder = "Relatorios_Gerados"
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        global_found_items = set() # Rastreia o que foi achado em QUALQUER planilha
        files_processed_count = 0

        try:
            # 1. Iterar sobre CADA arquivo carregado
            for filepath in self.filepaths:
                wb = load_workbook(filepath)
                
                # Iterar sobre CADA aba (sheet) dentro do arquivo
                for sheet in wb.worksheets:
                    # Busca Força Bruta
                    for row in sheet.iter_rows():
                        match_in_row = False
                        for cell in row:
                            if cell.value is not None:
                                cell_val = str(cell.value).strip()
                                if cell_val in self.scanned_codes:
                                    match_in_row = True
                                    global_found_items.add(cell_val)
                                    # Não damos break aqui, caso haja códigos duplicados na mesma linha
                        
                        if match_in_row:
                            for cell in row:
                                cell.fill = green_fill
                
                # Salva o arquivo modificado na pasta de saída
                filename = os.path.basename(filepath)
                save_path = os.path.join(output_folder, f"VERIFICADO_{filename}")
                wb.save(save_path)
                files_processed_count += 1

            # 2. Gerar Relatório de "Sobras" (O que foi bipado mas não estava nas listas)
            not_found_items = self.scanned_codes - global_found_items
            
            if not_found_items:
                wb_extras = Workbook()
                ws_extras = wb_extras.active
                ws_extras.title = "Itens Não Identificados"
                ws_extras.append(["Código Lido", "Status"]) # Cabeçalho
                
                for item in not_found_items:
                    ws_extras.append([item, "NÃO ENCONTRADO NAS PLANILHAS"])
                
                extras_path = os.path.join(output_folder, "RELATORIO_ITENS_NÃO_ENCONTRADOS.xlsx")
                wb_extras.save(extras_path)
                msg_extra = f"\n\nATENÇÃO: {len(not_found_items)} itens não foram achados.\nVeja o arquivo 'RELATORIO_ITENS_SOBRANDO.xlsx'."
            else:
                msg_extra = "\n\nSucesso Total: Todos os itens lidos foram encontrados!"

            messagebox.showinfo("Concluído", 
                                f"Processamento Finalizado!\n\n"
                                f"Arquivos processados: {files_processed_count}\n"
                                f"Salvos na pasta: /{output_folder}" + msg_extra)
            
            # Abre a pasta de resultados
            os.startfile(output_folder)

        except Exception as e:
            messagebox.showerror("Erro Crítico", f"Falha no processamento: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = MultiInventoryApp(root)
    root.mainloop()