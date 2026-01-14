import os
import zipfile
import io
import time
import threading
import requests
import unicodedata
import re
from flask import Flask, render_template, request, send_file, jsonify, session
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

# Local imports
from .models import db, User, Network, FileMetadata

app = Flask(__name__, static_folder=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static'))
app.secret_key = 'super_secret_key_sesi_sorocaba' # Change in production

# --- Database Config ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'database.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# --- File Storage Config ---
PROJECT_ROOT = os.path.abspath(os.path.join(BASE_DIR, '..'))
UPLOAD_FOLDER = os.path.join(PROJECT_ROOT, 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

SCANNED_DATA_FOLDER = os.path.join(UPLOAD_FOLDER, 'scanned_data')
os.makedirs(SCANNED_DATA_FOLDER, exist_ok=True)

REPORTS_FOLDER = os.path.join(PROJECT_ROOT, 'Relatorios_Gerados')
os.makedirs(REPORTS_FOLDER, exist_ok=True)

# Initialize DB
with app.app_context():
    db.create_all()

# --- Keep Alive System ---
def keep_alive_pinger():
    """Pings the server every 14 minutes to prevent sleep."""
    url = "http://127.0.0.1:8000/keep_alive" # Default Gunicorn port often 8000 or from env
    # In production (Render), we might need the public URL. 
    # But localhost ping might suffice if the server process itself stays active.
    # However, Render might kill if no external requests.
    # User requested "something to send requests".
    
    # Better approach for Render: Use the public URL if available, else localhost
    # Render sets RENDER_EXTERNAL_URL
    public_url = os.environ.get('RENDER_EXTERNAL_URL')
    target_url = public_url + "/keep_alive" if public_url else url

    print(f" * Keep-Alive Pinger Initialized. Target: {target_url}")
    
    while True:
        time.sleep(14 * 60) # 14 minutes
        try:
            print(" * Sending Keep-Alive Ping...")
            requests.get(target_url)
        except Exception as e:
            print(f" * Keep-Alive Ping Failed: {e}")

# Start Pinger in Background Thread
# Only start if not in debug/reloader mode to avoid duplicates
if not app.debug or os.environ.get("WERKZEUG_RUN_MAIN") == "true":
    t = threading.Thread(target=keep_alive_pinger, daemon=True)
    t.start()

@app.route('/keep_alive')
def keep_alive():
    return jsonify({"status": "alive", "timestamp": time.time()})

# --- Routes ---

@app.route('/get_active_cities', methods=['GET'])
def get_active_cities():
    # DISTINCT city from Network table
    cities = [r[0] for r in db.session.query(Network.city).distinct()]
    return jsonify({'cities': sorted(cities)})

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['POST'])
def login():
    data = request.json
    email = data.get('email')
    password = data.get('password')

    user = User.query.filter_by(email=email).first()

    if user and check_password_hash(user.password, password):
        if not user.is_admin:
             return jsonify({'error': 'Acesso negado. Apenas administradores.'}), 403
             
        session['user_id'] = user.id
        session['is_admin'] = True
        session['city'] = user.city
        
        # Super Admin Check
        is_super = (email == 'admin@123')
        session['is_super_admin'] = is_super 

        return jsonify({
            'message': 'Login realizado com sucesso', 
            'success': True,
            'city': user.city,
            'is_admin': True,
            'is_super_admin': is_super
        })
    else:
        return jsonify({'error': 'Credenciais inválidas', 'success': False}), 401

@app.route('/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'message': 'Logout realizado com sucesso'})

@app.route('/check_auth', methods=['GET'])
def check_auth():
    return jsonify({
        'is_admin': session.get('is_admin', False),
        'is_super_admin': session.get('is_super_admin', False),
        'city': session.get('city', None),
        'connected_network_id': session.get('connected_network_id', None),
        'connected_network_name': session.get('connected_network_name', None)
    })

# --- Network & Admin Management ---

@app.route('/register_admin', methods=['POST'])
def register_admin():
    data = request.json
    email = data.get('email')
    password = data.get('password')
    city = data.get('city')
    network_name = data.get('network_name')
    network_pass = data.get('network_password')
    
    if not all([email, password, city, network_name, network_pass]):
        return jsonify({'error': 'Preencha todos os campos'}), 400

    try:
        if User.query.filter_by(email=email).first():
             return jsonify({'error': 'E-mail já cadastrado'}), 400
        
        if Network.query.filter_by(name=network_name).first():
             return jsonify({'error': 'Nome da rede já existe.'}), 400

        hashed_pw = generate_password_hash(password)
        new_user = User(email=email, password=hashed_pw, city=city, is_admin=True)
        db.session.add(new_user)
        db.session.flush() # Get ID

        hashed_net_pw = generate_password_hash(network_pass)
        new_net = Network(name=network_name, password=hashed_net_pw, city=city, admin_id=new_user.id)
        db.session.add(new_net)
        
        db.session.commit()
        return jsonify({'message': 'Conta e Rede criadas com sucesso! Faça login.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/create_network', methods=['POST'])
def create_network():
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.json
    name = data.get('name')
    password = data.get('password')
    city = session.get('city') 
    
    if not name or not password: return jsonify({'error': 'Nome e Senha obrigatórios'}), 400
    
    try:
        if Network.query.filter_by(name=name).first():
             return jsonify({'error': 'Nome de rede já existe'}), 400

        hashed = generate_password_hash(password)
        # Fix: ensure user_id is int
        uid = int(session.get('user_id'))
        
        new_net = Network(name=name, password=hashed, city=city, admin_id=uid)
        db.session.add(new_net)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/delete_network', methods=['POST'])
def delete_network():
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    data = request.json
    net_id = data.get('id')
    
    net = Network.query.get(net_id)
    if not net:
        return jsonify({'error': 'Rede não encontrada'}), 404

    # Verify ownership
    if net.admin_id != int(session.get('user_id')):
        if not session.get('is_super_admin'):
             return jsonify({'error': 'Acesso negado'}), 403
        
    db.session.delete(net)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/get_my_networks', methods=['GET'])
def get_my_networks():
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    
    uid = int(session.get('user_id'))
    nets = Network.query.filter_by(admin_id=uid).all()
    return jsonify({'networks': [{'id': n.id, 'name': n.name} for n in nets]})

@app.route('/get_networks', methods=['GET'])
def get_networks():
    city = request.args.get('city')
    if not city: return jsonify({'networks': []})
    
    nets = Network.query.filter_by(city=city).all()
    results = []
    for n in nets:
        owner = User.query.get(n.admin_id)
        owner_email = owner.email if owner else 'Unknown'
        results.append({
            'id': n.id,
            'name': n.name,
            'owner': owner_email
        })
    return jsonify({'networks': results})

@app.route('/join_network', methods=['POST'])
def join_network():
    data = request.json
    try:
        network_id = int(data.get('network_id'))
    except:
        return jsonify({'error': 'ID inválido'}), 400
        
    password = data.get('password')
    
    network = Network.query.get(network_id)
    
    # Super Admin Bypass or Password Check
    if network:
        # Check bypass BEFORE clearing session
        is_super = session.get('is_super_admin')
        
        if is_super or check_password_hash(network.password, password):
            session.clear()
            session['connected_network_id'] = network.id
            session['connected_network_name'] = network.name
            session['city'] = network.city
            session['is_admin'] = False
            # Restore Super Admin if applicable
            if is_super: session['is_super_admin'] = True
            
            return jsonify({'success': True, 'message': f'Conectado à rede {network.name}'})
        else:
             return jsonify({'error': 'Senha da rede incorreta.'}), 401
    else:
        return jsonify({'error': 'Rede não encontrada.'}), 404

# --- File Management (Local FS + SQL Metadata) ---

@app.route('/upload_master', methods=['POST'])
def upload_master():
    if not session.get('is_admin'):
        return jsonify({'error': 'Acesso negado.'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    if file and file.filename.lower().endswith('.xlsx'):
        filename = file.filename
        
        # Save to Disk
        safe_name = re.sub(r'[^a-zA-Z0-9_.-]', '_', filename)
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
        file.save(save_path)
        
        # Save Metadata
        user_id = int(session.get('user_id'))
        network_id = request.form.get('network_id')
        if network_id: network_id = int(network_id)
        
        new_file = FileMetadata(
            filename=filename,
            filepath=safe_name,
            type='master_spreadsheet',
            user_id=user_id,
            network_id=network_id
        )
        db.session.add(new_file)
        db.session.commit()
        
        return jsonify({'message': f'Planilha "{filename}" carregada com sucesso!'})
    
    return jsonify({'error': 'Formato inválido. Apenas .xlsx'}), 400

@app.route('/list_masters', methods=['GET'])
def list_masters():
    user_id = session.get('user_id')
    connected_net_id = session.get('connected_network_id')
    is_super = session.get('is_super_admin', False)
    
    query = FileMetadata.query.filter_by(type='master_spreadsheet')
    
    # Simple permissions
    if is_super:
        pass
    elif user_id:
        network_id = request.args.get('network_id')
        if network_id:
            query = query.filter_by(network_id=int(network_id))
        else:
            query = query.filter_by(user_id=int(user_id))
    elif connected_net_id:
        # Public: files from this network OR basic user files? 
        # Keeping logic: show files for this network
        query = query.filter((FileMetadata.network_id == int(connected_net_id)))
    else:
        return jsonify({'masters': []})

    files = query.all()
    return jsonify({'masters': sorted([f.filename for f in files])})

@app.route('/delete_master', methods=['POST'])
def delete_master():
    if not session.get('is_admin'): return jsonify({'error': 'Acesso negado.'}), 403

    filename = request.json.get('filename')
    f_meta = FileMetadata.query.filter_by(filename=filename, type='master_spreadsheet').first()
    
    if not f_meta: return jsonify({'error': 'Arquivo não encontrado'}), 404
        
    if not session.get('is_super_admin'):
        if f_meta.user_id != int(session.get('user_id')):
            return jsonify({'error': 'Permissão negada'}), 403
            
    try:
        # Remove from Disk
        full_path = os.path.join(app.config['UPLOAD_FOLDER'], f_meta.filepath)
        if os.path.exists(full_path):
            os.remove(full_path)
            
        # Remove DB
        db.session.delete(f_meta)
        db.session.commit()
        return jsonify({'message': 'Removido com sucesso'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get_master/<filename>', methods=['GET'])
def get_master(filename):
    if not session.get('is_admin'): return jsonify({'error': 'Acesso negado.'}), 403
    
    f_meta = FileMetadata.query.filter_by(filename=filename).first()
    if not f_meta: return jsonify({'error': 'Arquivo não encontrado'}), 404
    
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], f_meta.filepath), download_name=filename)

# --- Verification & Logic ---

@app.route('/list_reports', methods=['GET'])
def list_reports():
    query = FileMetadata.query.filter_by(type='audit_report')
    
    net_id = session.get('connected_network_id') or request.args.get('network_id')
    user_id = session.get('user_id')
    
    if session.get('is_super_admin'):
        pass # See ALL
    elif session.get('is_admin'):
        # Admin sees reports from ALL networks they manage
        if user_id:
            my_nets = Network.query.filter_by(admin_id=int(user_id)).with_entities(Network.id).all()
            my_net_ids = [n.id for n in my_nets]
            if my_net_ids:
                query = query.filter(FileMetadata.network_id.in_(my_net_ids))
            else:
                 query = query.filter_by(user_id=int(user_id)) # Fallback if no networks
    elif net_id:
        query = query.filter_by(network_id=int(net_id))
    else:
        return jsonify({'reports': []})
        
    files = query.all()
    # Return more info for admin visibility
    return jsonify({'reports': [{'filename': f.filename, 'network_id': f.network_id} for f in files]})

@app.route('/delete_report', methods=['POST'])
def delete_report():
    if not session.get('is_admin'): return jsonify({'error': 'Unauthorized'}), 403
    filename = request.json.get('filename')
    f_meta = FileMetadata.query.filter_by(filename=filename, type='audit_report').first()
    
    if f_meta:
        path = os.path.join(REPORTS_FOLDER, f_meta.filepath)
        if os.path.exists(path): os.remove(path)
        db.session.delete(f_meta)
        db.session.commit()
    return jsonify({'success': True})

@app.route('/get_report/<path:filename>', methods=['GET'])
def get_report(filename):
    if not session.get('is_admin') and not session.get('connected_network_id'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    report_path = os.path.join(REPORTS_FOLDER, filename)
    if os.path.exists(report_path):
        return send_file(report_path, as_attachment=True, download_name=filename)
        
    f_meta = FileMetadata.query.filter_by(filename=filename).first()
    if f_meta:
        path = os.path.join(app.config['UPLOAD_FOLDER'], f_meta.filepath)
        if os.path.exists(path):
            return send_file(path, as_attachment=True, download_name=filename)
            
    return jsonify({'error': 'Arquivo não encontrado'}), 404

@app.route('/get_rooms', methods=['POST'])
def get_rooms():
    data = request.json
    selected_files = data.get('filenames', [])
    all_rooms = []

    for filename in selected_files:
        f_meta = FileMetadata.query.filter_by(filename=filename).first()
        if not f_meta: continue
        
        path = os.path.join(app.config['UPLOAD_FOLDER'], f_meta.filepath)
        if not os.path.exists(path): continue
        
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                
                sheet_rooms = []
                
                # Scan for Header Row
                for r_idx, row in enumerate(rows):
                    # Check if this row looks like the specific header row
                    # We look for "Localização" and "Denominação"
                    row_str = [str(c).strip().lower() for c in row if c]
                    
                    if any("localização" in s for s in row_str):
                        # Found a potential header row.
                        # Now try to identify column indices
                        loc_idx = -1
                        denom_idx = -1
                        inv_idx = -1
                        
                        for c_idx, cell in enumerate(row):
                            val = str(cell).strip()
                            val_lower = val.lower()
                            
                            if "localização" in val_lower:
                                loc_idx = c_idx
                            elif "denominação" in val_lower and "imobilizado" not in val_lower:
                                # Avoid "Denominação do imobilizado" which is the items list
                                denom_idx = c_idx
                            elif "nº invent" in val_lower or "n° invent" in val_lower:
                                # Grab Inventory Number if available
                                inv_idx = c_idx
                        
                        # If we found at least Localização, check the NEXT row for data
                        if loc_idx != -1 and r_idx + 1 < len(rows):
                            data_row = rows[r_idx + 1]
                            
                            loc_val = str(data_row[loc_idx]).strip() if loc_idx < len(data_row) and data_row[loc_idx] else ""
                            denom_val = str(data_row[denom_idx]).strip() if denom_idx != -1 and denom_idx < len(data_row) and data_row[denom_idx] else ""
                            inv_val = str(data_row[inv_idx]).strip() if inv_idx != -1 and inv_idx < len(data_row) and data_row[inv_idx] else ""
                            
                            # Construct Display Name
                            # Filter empty parts
                            parts = [p for p in [loc_val, denom_val, inv_val] if p and p != "None"]
                            if parts:
                                full_name = " - ".join(parts)
                                # Room ID: Use Sheet Name + Localização (or just Sheet Name if unique per sheet)
                                # To be safe and robust:
                                room_id = f"{sheet_name}::{full_name}"
                                sheet_rooms.append({'id': room_id, 'name': full_name, 'source': filename, 'type': 'sliced'})
                                
                                # Assume only one main header per sheet for this specific format? 
                                # Or continue scanning?
                                # If we found a valid room line, unlikely to change format in same sheet.
                                # But let's continue scanning just in case multiple rooms are listed (unlikely based on images).
                                break 

                if sheet_rooms:
                    all_rooms.extend(sheet_rooms)
                else:
                    # Fallback: If no headers found, do NOT add generic sheet names "Table X"
                    # User requested to remove "Table 1, Table 2..." garbage.
                    # Only add if it looks like a meaningful sheet? 
                    # For now, suppressing fallback as per user request to clean up list.
                    pass
                    
            wb.close()
        except: pass

    return jsonify({'rooms': all_rooms})

@app.route('/verify', methods=['POST'])
def verify():
    data = request.json
    analyst_name = data.get('analyst_name', 'Analista')
    selected_room = data.get('room_name')
    source_file = data.get('source_file')
    scanned_codes_raw = data.get('scanned_codes', '')
    
    if not source_file: return jsonify({'error': 'Arquivo fonte não identificado'}), 400

    # Clean Scanned Codes
    scanned_codes = set()
    for line in scanned_codes_raw.splitlines():
        c = line.strip()
        if c: scanned_codes.add(c)
    
    f_meta = FileMetadata.query.filter_by(filename=source_file).first()
    if not f_meta: return jsonify({'error': 'Arquivo não encontrado db'}), 404
    
    path = os.path.join(app.config['UPLOAD_FOLDER'], f_meta.filepath)
    if not os.path.exists(path): return jsonify({'error': 'Arquivo físico não encontrado'}), 404

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    try:
        # 1. Parse Expected Items from Sheet
        wb = load_workbook(path, read_only=True, data_only=True)
        # Parse Room ID to get Sheet Name and Row Offset if sliced
        # Format: "SheetName::Localização - Denom..."
        is_sliced = "::" in selected_room
        target_sheet_name = selected_room.split("::")[0] if is_sliced else selected_room
        
        if target_sheet_name not in wb.sheetnames: return jsonify({'error': 'Aba não encontrada'}), 400
        ws = wb[target_sheet_name]
        
        rows = list(ws.iter_rows(values_only=True))
        
        # Identify Columns (Dynamic like get_rooms)
        inv_idx = -1
        desc_idx = -1
        expected_items = {} # Map Code -> Row Data (or Description)
        
        # Scan header to find columns
        header_row_idx = -1
        for r_idx, row in enumerate(rows):
            row_str = [str(c).strip().lower() for c in row if c]
            if any("nº invent" in s or "n° invent" in s for s in row_str):
                 header_row_idx = r_idx
                 # Map cols
                 for c_idx, cell in enumerate(row):
                     val = str(cell).strip().lower()
                     if "nº invent" in val or "n° invent" in val: inv_idx = c_idx
                     elif "denominação" in val: desc_idx = c_idx
                 break
        
        # Extract Expected
        if header_row_idx != -1 and inv_idx != -1:
            for r_idx in range(header_row_idx + 1, len(rows)):
                row = rows[r_idx]
                if inv_idx < len(row) and row[inv_idx]:
                    code = str(row[inv_idx]).strip()
                    # Filter: If "sliced", verify if this row matches the specific room?
                    # The user selected a "Sliced" room (specific header value).
                    # But the current get_rooms implementation just gives us lines.
                    # Complex Logic: If the sheet has MULTIPLE rooms, we need to filter only items for THAT room.
                    # Current `get_rooms` finds the header "Localização". 
                    # If this is a master sheet with one room per sheet, we take all.
                    # If it's a huge sheet with many rooms... we need to filter by "Localização" column if it exists.
                    # Let's assume for now we take all items in that sheet OR try to match the room name.
                    # Given the "Sliced" logic in `get_rooms` uses the header VALUE to name the room...
                    # We should check if the row matches that value.
                    
                    # Refined Logic:
                    # If is_sliced, selected_room contains "Sheet::Loc - Denom - Inv".
                    # We can't easily filter back without knowing which column is Localização.
                    # Let's try to grab Description for context.
                    desc = str(row[desc_idx]).strip() if desc_idx != -1 and desc_idx < len(row) else "Item"
                    expected_items[code] = {'desc': desc, 'row_data': row}
        
        wb.close()
        
        # 2. Compare
        verified_codes = []
        missing_codes = []
        extra_codes = []
        
        # Verified & Missing
        for code, info in expected_items.items():
            if code in scanned_codes:
                verified_codes.append({'code': code, 'desc': info['desc'], 'status': 'Encontrado'})
            else:
                missing_codes.append({'code': code, 'desc': info['desc'], 'status': 'Faltante'})
        
        # Extra
        for code in scanned_codes:
            if code not in expected_items:
                extra_codes.append({'code': code, 'desc': 'Não consta na planilha', 'status': 'Sobras'})
                
        # 3. Generate 3 Excel Files
        from openpyxl import Workbook
        
        def save_excel(data, filename, title):
            wb_new = Workbook()
            ws_new = wb_new.active
            ws_new.title = title
            ws_new.append(["Código", "Descrição", "Status"])
            for item in data:
                ws_new.append([item['code'], item['desc'], item['status']])
            path = os.path.join(app.config['UPLOAD_FOLDER'], filename) # Temp save
            wb_new.save(path)
            return path

        files_to_zip = []
        
        # File 1: Analisados (Verified)
        f1 = f"Conferidos_{analyst_name}_{timestamp}.xlsx"
        p1 = save_excel(verified_codes, f1, "Conferidos")
        files_to_zip.append((p1, f1))
        
        # File 2: Deveriam ter sido encontrados (Missing)
        f2 = f"Faltantes_{analyst_name}_{timestamp}.xlsx"
        p2 = save_excel(missing_codes, f2, "Faltantes")
        files_to_zip.append((p2, f2))
        
        # File 3: Não encontrados/Sobras (Extra)
        f3 = f"Sobras_{analyst_name}_{timestamp}.xlsx"
        p3 = save_excel(extra_codes, f3, "Sobras")
        files_to_zip.append((p3, f3))
        
        # 4. ZIP Them
        zip_filename = f"Auditoria_{analyst_name}_{timestamp}.zip"
        zip_path = os.path.join(REPORTS_FOLDER, zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for file_path, arcname in files_to_zip:
                zipf.write(file_path, arcname)
                
        # Cleanup temp xl files
        for file_path, _ in files_to_zip:
            if os.path.exists(file_path): os.remove(file_path)
            
        # 5. Metadata
        net_id = session.get('connected_network_id')
        user_id = session.get('user_id')
        
        new_rep = FileMetadata(
            filename=zip_filename,
            filepath=zip_filename,
            type='audit_report',
            user_id=int(user_id) if user_id else None,
            network_id=int(net_id) if net_id else None
        )
        db.session.add(new_rep)
        db.session.commit()

        return jsonify({
            'success': True,
            'download_url': f'/get_report/{zip_filename}'
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# --- Admin Users ---
@app.route('/admin/users', methods=['GET'])
def list_all_users():
    if not session.get('is_super_admin'): return jsonify({'error': 'Unauthorized'}), 403
    users = User.query.all()
    return jsonify({'users': [{'id': u.id, 'email': u.email, 'city': u.city, 'is_admin': u.is_admin} for u in users]})

@app.route('/admin/users/<int:user_id>', methods=['DELETE'])
def delete_user_account(user_id):
    if not session.get('is_super_admin'): return jsonify({'error': 'Unauthorized'}), 403
    u = User.query.get(user_id)
    if u:
        db.session.delete(u)
        db.session.commit()
    return jsonify({'success': True})

if __name__ == '__main__':
    app.run(debug=True)
